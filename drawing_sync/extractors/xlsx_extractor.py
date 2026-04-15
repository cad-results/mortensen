"""XLSX Schedule/BOM Extractor.

Extracts component data from Excel schedule files:
- Drawing indices
- Junction box BOMs
- Panel elevation schedules
- Cable/fiber schedules
- P&C abbreviation tables

Performs exhaustive regex extraction (parity with PDF/DXF extractors)
and full BOM line-item extraction for BOM files.
"""

import re
import os
from typing import Optional

from ..models import (
    Component, ComponentType, ComponentValue, DrawingData, TextLabel,
)


class XLSXExtractor:
    """Extracts component and schedule data from XLSX files."""

    def __init__(self):
        self._openpyxl = None
        self._xlrd = None
        self._current_is_bom = False

    def _get_openpyxl(self):
        if self._openpyxl is None:
            import openpyxl
            self._openpyxl = openpyxl
        return self._openpyxl

    def _get_xlrd(self):
        if self._xlrd is None:
            try:
                import xlrd
                self._xlrd = xlrd
            except ImportError:
                self._xlrd = False  # Mark as unavailable
        return self._xlrd

    # ─── BOM detection ────────────────────────────────────────────────

    @staticmethod
    def _is_bom_file(filename: str) -> bool:
        """Check if this file is a BOM based on filename."""
        return "BOM" in filename.upper()

    @staticmethod
    def _detect_bom_sheet(header_cells: list, sheet_name: str):
        """Detect if a sheet is a BOM sheet and what type.

        Returns (is_bom, bom_type) where bom_type is:
          "material"    — ITEM#, QTY, CATALOG#, MATERIAL columns
          "device_list" — ITEM#, DEVICE#, QTY, CATALOG#, MATERIAL columns
          "nameplate"   — NAMEPLATE SCHEDULE header
          ""            — not a BOM sheet
        """
        upper = [c.upper() for c in header_cells]
        joined = " ".join(upper)

        # Nameplate sheets: first cell or sheet name indicates it
        if "NAMEPLATE SCHEDULE" in joined or "NAMEPLATE" in sheet_name.upper():
            return True, "nameplate"

        has_item = any("ITEM" in c for c in upper)
        has_qty = any("QTY" in c or "QUANTITY" in c for c in upper)
        has_material = any("MATERIAL" in c for c in upper)
        has_catalog = any("CATALOG" in c for c in upper)
        has_device = any("DEVICE" in c for c in upper)

        if has_item and has_qty and has_device:
            return True, "device_list"
        if has_item and has_qty and (has_material or has_catalog):
            return True, "material"

        return False, ""

    # ─── BOM material classification ──────────────────────────────────

    @staticmethod
    def _classify_bom_material(description: str, qty_str: str,
                               catalog_number: str) -> ComponentType:
        """Classify a BOM material item into a ComponentType."""
        desc = description.upper()
        qty_upper = qty_str.upper()

        # Consumables first (LOT quantity overrides everything)
        if "LOT" in qty_upper:
            return ComponentType.CONSUMABLE

        # Consumable keywords
        consumable_kw = [
            "CRIMP TERMINAL", "THHN", " SIS,", " SIS ", "WIRE ",
            "SPRING NUT", "STRUT", "ROD WITH MOUNTING",
            "CHANNEL", "MINI STRUT",
        ]
        if any(kw in desc for kw in consumable_kw):
            return ComponentType.CONSUMABLE

        # Terminal blocks
        if "TERMINAL BLOCK" in desc or "BARRIER" in desc:
            return ComponentType.TERMINAL_BLOCK

        # Fuses
        if "FUSE" in desc or (catalog_number and "BAF" in catalog_number.upper()):
            return ComponentType.FUSE

        # Panel for enclosure (check before ENCLOSURE to avoid false match)
        if "PANEL FOR ENCLOSURE" in desc or "PANEL FOR" in desc:
            return ComponentType.PANEL

        # Enclosures / junction boxes (including typo "ENCLOSRE")
        if any(kw in desc for kw in ["ENCLOSURE", "ENCLOSRE", "CABINET"]):
            return ComponentType.JUNCTION_BOX

        # Ground bar
        if "GROUND BAR" in desc or "GROUND BUS" in desc:
            return ComponentType.GROUND

        # DIN rail
        if "DIN RAIL" in desc:
            return ComponentType.BOM_ITEM

        # Cable/coax/fiber
        if any(kw in desc for kw in ["CABLE", "COAX", "RG-58", "FIBER", "BNC"]):
            return ComponentType.CABLE

        # Resistor / terminator
        if "RESISTOR" in desc or "TERMINATOR" in desc:
            return ComponentType.BOM_ITEM

        # Tee / connector
        if "TEE" in desc or "T-CONN" in desc or "CONNECTOR" in desc:
            return ComponentType.BOM_ITEM

        # Power supply
        if "POWER SUPPLY" in desc:
            return ComponentType.POWER_SUPPLY

        return ComponentType.BOM_ITEM

    # ─── Component ID generation ──────────────────────────────────────

    @staticmethod
    def _generate_bom_component_id(item_num: str, device_name: str,
                                   catalog_number: str, description: str,
                                   existing_ids: set) -> str:
        """Generate a unique component ID for a BOM line item."""
        # Priority 1: Device name (IRIG-B style)
        if device_name:
            base_id = device_name.strip()
            if base_id not in existing_ids:
                return base_id
            # Append item number for dedup
            dedup_id = f"{base_id}-{item_num}"
            if dedup_id not in existing_ids:
                return dedup_id
            return f"{base_id}:{item_num}"

        # Priority 2: Catalog number
        if catalog_number:
            base_id = catalog_number.strip()
            if base_id not in existing_ids:
                return base_id
            dedup_id = f"{base_id}:{item_num}"
            return dedup_id

        # Priority 3: Fallback
        return f"BOM-ITEM-{item_num}"

    # ─── Main extract entry point ─────────────────────────────────────

    def extract(self, file_path: str) -> DrawingData:
        """Extract data from an XLSX/XLS schedule file."""
        filename = os.path.basename(file_path)
        drawing_id = os.path.splitext(filename)[0]
        ext = os.path.splitext(filename)[1].lower()

        drawing = DrawingData(
            drawing_id=drawing_id,
            file_path=file_path,
            file_type="xlsx" if ext != ".xls" else "xls",
        )

        self._current_is_bom = self._is_bom_file(filename)

        if ext == ".xls":
            return self._extract_xls(file_path, drawing)

        try:
            openpyxl = self._get_openpyxl()
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            drawing.notes.append(f"XLSX read error: {str(e)}")
            return drawing

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self._extract_sheet(ws, sheet_name, drawing)

        wb.close()
        return drawing

    def _extract_xls(self, file_path: str, drawing: DrawingData) -> DrawingData:
        """Extract data from an old-format .xls file using xlrd."""
        xlrd = self._get_xlrd()
        if not xlrd:
            drawing.notes.append(
                "XLS read skipped: xlrd not installed. "
                "Install with: pip install xlrd"
            )
            return drawing

        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as e:
            drawing.notes.append(f"XLS read error: {str(e)}")
            return drawing

        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for row_idx in range(ws.nrows):
                cells = [
                    str(ws.cell_value(row_idx, c)).strip()
                    for c in range(ws.ncols)
                ]
                rows.append(cells)
            self._extract_sheet_from_rows(rows, sheet_name, drawing)

        return drawing

    # ─── Sheet extraction ─────────────────────────────────────────────

    def _extract_sheet(self, ws, sheet_name: str, drawing: DrawingData):
        """Extract data from an openpyxl worksheet."""
        # Pre-read all rows
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            rows.append(cells)

        self._extract_sheet_from_rows(rows, sheet_name, drawing)

    def _extract_sheet_from_rows(self, rows: list, sheet_name: str,
                                 drawing: DrawingData):
        """Core sheet extraction from pre-read row data.

        Works for both openpyxl (.xlsx) and xlrd (.xls) sources.
        """
        if not rows:
            return

        # Find header row — prefer rows with recognizable column headers
        # (ITEM, QTY, MATERIAL, etc.) over title rows
        header_row = []
        header_row_idx = 0
        best_header_idx = None
        best_header_score = 0
        for idx, row in enumerate(rows[:5]):
            if not any(row):
                continue
            upper = " ".join(row).upper()
            score = sum(1 for kw in [
                "ITEM", "QTY", "QUANTITY", "MATERIAL", "CATALOG",
                "DESCRIPTION", "DEVICE", "SIZE", "DRAWING", "DWG",
            ] if kw in upper)
            if score > best_header_score:
                best_header_score = score
                best_header_idx = idx
            if header_row_idx == 0 and not header_row:
                header_row = row
                header_row_idx = idx

        if best_header_idx is not None and best_header_score >= 2:
            header_row = rows[best_header_idx]
            header_row_idx = best_header_idx

        # ── BOM-specific extraction (additive, runs first) ──
        if self._current_is_bom:
            is_bom_sheet, bom_type = self._detect_bom_sheet(
                header_row, sheet_name,
            )
            if is_bom_sheet:
                if bom_type == "nameplate":
                    self._extract_nameplate_sheet(rows, sheet_name, drawing)
                else:
                    self._extract_bom_sheet(
                        rows, header_row_idx, sheet_name, drawing,
                    )

        # ── Existing extraction (always runs) ──
        for row_idx, cells in enumerate(rows):
            row_text = " ".join(cells)

            # Store all text
            drawing.raw_text += row_text + "\n"

            # Look for component references (exhaustive regex)
            self._extract_components_from_row(
                cells, header_row, drawing,
            )

            # Store as labels
            for col_idx, cell in enumerate(cells):
                if cell:
                    drawing.all_labels.append(TextLabel(
                        text=cell,
                        x=col_idx,
                        y=row_idx + 1,
                        category=self._categorize_cell(
                            cell, header_row, col_idx,
                        ),
                    ))

    # ─── BOM sheet extraction ─────────────────────────────────────────

    def _extract_bom_sheet(self, rows: list, header_idx: int,
                           sheet_name: str, drawing: DrawingData):
        """Extract all BOM line items from a material/device-list sheet."""
        if header_idx >= len(rows):
            return

        header = rows[header_idx]
        col_map = self._build_column_map(header)

        existing_ids = set(drawing.components.keys())

        for row_idx in range(header_idx + 1, len(rows)):
            cells = rows[row_idx]
            if not any(cells):
                continue

            # Extract all named fields
            item_num = self._get_col(cells, col_map, "item", "").strip()
            qty = self._get_col(cells, col_map, "qty", "").strip()
            catalog = self._get_col(cells, col_map, "catalog", "").strip()
            material = self._get_col(cells, col_map, "material", "").strip()
            device = self._get_col(cells, col_map, "device", "").strip()

            # Skip rows that look like sub-headers or empty data
            if not item_num and not material and not catalog and not device:
                continue

            # Use row index as fallback item number
            if not item_num:
                item_num = str(row_idx)

            # Generate component ID
            comp_id = self._generate_bom_component_id(
                item_num, device, catalog, material, existing_ids,
            )
            existing_ids.add(comp_id)

            # Classify component type
            comp_type = self._classify_bom_material(material, qty, catalog)

            # Build full attributes dict from ALL columns
            attributes = {"bom_source": "material", "sheet_name": sheet_name}
            for col_name, col_idx in col_map.items():
                if col_idx < len(cells) and cells[col_idx]:
                    attributes[col_name] = cells[col_idx]
            # Also store normalized fields
            attributes["item_number"] = item_num
            if qty:
                attributes["qty"] = qty
            if catalog:
                attributes["catalog_number"] = catalog
            if device:
                attributes["device_name"] = device

            # Create or merge component
            if comp_id in drawing.components:
                # Merge attributes into existing component
                drawing.components[comp_id].attributes.update(attributes)
                if material and not drawing.components[comp_id].description:
                    drawing.components[comp_id].description = material
            else:
                comp = Component(
                    component_id=comp_id,
                    component_type=comp_type,
                    description=material,
                    attributes=attributes,
                )
                # Extract embedded electrical values from description
                self._extract_values_from_text(material, comp)
                drawing.components[comp_id] = comp

    def _extract_nameplate_sheet(self, rows: list, sheet_name: str,
                                 drawing: DrawingData):
        """Extract nameplate sub-components from a NAMEPLATE sheet."""
        # Find the actual header row (contains ITEM#, SIZE, etc.)
        header_idx = None
        for idx, row in enumerate(rows[:5]):
            upper = " ".join(row).upper()
            if "ITEM" in upper and ("SIZE" in upper or "LINE" in upper):
                header_idx = idx
                break

        if header_idx is None:
            # Store nameplate info text as notes
            for row in rows:
                text = " ".join(row).strip()
                if text:
                    drawing.notes.append(text)
            return

        header = rows[header_idx]
        col_map = self._build_column_map(header)

        # Store the material spec line (usually row 2) as a note
        for idx in range(header_idx):
            text = " ".join(rows[idx]).strip()
            if text and "MATERIAL:" in text.upper():
                drawing.notes.append(text)

        # Determine sheet prefix for multi-sheet BOMs (YARD LIGHT: LT1, LT2...)
        sheet_prefix = ""
        upper_name = sheet_name.upper().replace("NAMEPLATE", "").strip()
        if upper_name and upper_name not in ("", "SCHEDULE"):
            sheet_prefix = upper_name.rstrip()

        for row_idx in range(header_idx + 1, len(rows)):
            cells = rows[row_idx]
            if not any(cells):
                continue

            line1 = self._get_col(cells, col_map, "1st_line", "").strip()
            line2 = self._get_col(cells, col_map, "2nd_line", "").strip()
            item_num = self._get_col(cells, col_map, "item", "").strip()
            size = self._get_col(cells, col_map, "size", "").strip()
            letter_size = self._get_col(
                cells, col_map, "letter_size", "",
            ).strip()
            qty = self._get_col(cells, col_map, "qty", "").strip()

            if not line1:
                continue

            # Determine component ID from 1st LINE
            comp_id = line1.strip()

            # For multi-sheet BOMs, prefix with sheet context
            if sheet_prefix and comp_id in ("TB", "TB1"):
                comp_id = f"{sheet_prefix}-{comp_id}"
            elif sheet_prefix and not any(
                comp_id.startswith(p) for p in (
                    "34.5kV", "138kV", "NORTH", "SOUTH", "EAST",
                    "H-FRAME", "NE ", "SW ", "NW ", "SE ",
                )
            ):
                # Only prefix sub-component designations, not title nameplates
                if len(comp_id) <= 6 and not comp_id[0].isdigit():
                    comp_id = f"{sheet_prefix}-{comp_id}"

            # Build attributes
            attributes = {
                "bom_source": "nameplate",
                "sheet_name": sheet_name,
                "nameplate_line1": line1,
            }
            if line2:
                attributes["nameplate_line2"] = line2
            if size:
                attributes["nameplate_size"] = size
            if letter_size:
                attributes["nameplate_letter_size"] = letter_size
            if qty:
                attributes["nameplate_qty"] = qty
            if item_num:
                attributes["item_number"] = item_num

            # Store all raw columns too
            for col_name, col_idx in col_map.items():
                if col_idx < len(cells) and cells[col_idx]:
                    key = f"nameplate_{col_name}"
                    if key not in attributes:
                        attributes[key] = cells[col_idx]

            # Check if component already exists (from device-pattern or BOM)
            if comp_id in drawing.components:
                drawing.components[comp_id].attributes.update(attributes)
                continue

            # Infer type from designation prefix
            comp_type = self._infer_type_from_designation(comp_id, line2)

            comp = Component(
                component_id=comp_id,
                component_type=comp_type,
                description=line2 if line2 else line1,
                attributes=attributes,
            )
            drawing.components[comp_id] = comp

    # ─── Exhaustive regex extraction (parity with PDF/DXF) ────────────

    def _extract_components_from_row(
        self, cells: list, header: list, drawing: DrawingData,
    ):
        """Extract component info from a row of cells using ALL patterns."""
        row_text = " ".join(cells)

        from .pdf_extractor import (
            DEVICE_PATTERNS, RELAY_PATTERNS, INSTRUMENT_TX_PATTERNS,
            OTHER_PATTERNS, EXTENDED_PATTERNS,
            DPAC_PATTERN, OUTPUT_PATTERN, INPUT_PATTERN, CIRCUIT_PATTERN,
            DFR_NORM_PATTERN, CM_PATTERN, VM_PATTERN, CMET_PATTERN,
            FPP_NORM_PATTERN, HS_PATTERN, TEE_PATTERN, AUX_CONTACT_PATTERN,
            WSL_PATTERN, PS_PATTERN, RP_PATTERN, ETM_PATTERN,
            LTC_PATTERN, SST_PATTERN, AC_PANEL_PATTERN, NAMED_FUSE_PATTERN,
            BESS_PATTERN, IRIG_PATTERN, MULTIPAIR_CABLE_PATTERN,
            TC_PATTERN, LOR_PATTERN, VDC_PATTERN, COMM_MODULE_PATTERN,
            CT_CLASS_PATTERN, STANDALONE_TB_PATTERN, STANDALONE_TS_PATTERN,
            RTAC_PATTERN, CLOCK_PATTERN, PDC_PATTERN,
            CISCO_PATTERN, RTR_PATTERN, BAF_PATTERN, POWERSTRIP_PATTERN,
            ATS_PATTERN, RTU_PATTERN, HVAC_PATTERN, BP_PATTERN,
            EMS_PATTERN, BATTERY_EQUIP_PATTERN,
            UPS_PATTERN, MCC_PATTERN, SWGR_PATTERN, VFD_PATTERN,
            PLC_PATTERN, RECTIFIER_PATTERN, MOV_PATTERN, ARRESTER_PATTERN,
            SPD_PATTERN, REGULATOR_PATTERN, GENERATOR_PATTERN,
            VOLTAGE_PATTERN, CURRENT_PATTERN, IMPEDANCE_PATTERN,
            MVA_PATTERN, RATIO_PATTERN,
            CABLE_PATTERN, DRAWING_REF_PATTERN,
            _RELAY_PREFIX_MAP,
        )

        # ── IEEE/ANSI device patterns ──
        for comp_type, pattern in DEVICE_PATTERNS.items():
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1)
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=comp_type,
                    )
                self._add_row_attributes(
                    drawing.components[comp_id], cells, header,
                )

        # ── Relay model patterns ──
        for pat_idx, pattern in enumerate(RELAY_PATTERNS):
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1)
                # Apply manufacturer prefix if needed
                if pat_idx in _RELAY_PREFIX_MAP:
                    comp_id = _RELAY_PREFIX_MAP[pat_idx] + comp_id
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=ComponentType.RELAY,
                    )
                self._add_row_attributes(
                    drawing.components[comp_id], cells, header,
                )

        # ── Instrument transformer patterns ──
        for comp_type, pattern in INSTRUMENT_TX_PATTERNS.items():
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1)
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=comp_type,
                    )
                self._add_row_attributes(
                    drawing.components[comp_id], cells, header,
                )

        # ── OTHER_PATTERNS (fuses, NGR, switches, panels) ──
        for comp_type, pattern in OTHER_PATTERNS.items():
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1).strip()
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=comp_type,
                    )
                self._add_row_attributes(
                    drawing.components[comp_id], cells, header,
                )

        # ── DPAC controllers (normalize DPAC1 → DPAC-1) ──
        for match in re.finditer(DPAC_PATTERN, row_text):
            comp_id = f"DPAC-{match.group(2)}"
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RELAY,
                    description=f"DPAC controller {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        # ── EXTENDED_PATTERNS (SEL communication cables) ──
        for comp_type, pattern in EXTENDED_PATTERNS.items():
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1).strip()
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=comp_type,
                    )
                self._add_row_attributes(
                    drawing.components[comp_id], cells, header,
                )

        # ── Output/Input contacts ──
        for match in re.finditer(OUTPUT_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RELAY,
                    description=f"Relay output contact {comp_id}",
                )

        for match in re.finditer(INPUT_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RELAY,
                    description=f"Relay input contact {comp_id}",
                )

        # ── Circuit identifiers ──
        for match in re.finditer(CIRCUIT_PATTERN, row_text):
            comp_id = f"CIRCUIT-{match.group(1)}"
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.BREAKER,
                    description=f"DC panel circuit breaker #{match.group(1)}",
                )

        # ── DFR (normalize DFR1 → DFR-1) ──
        for match in re.finditer(DFR_NORM_PATTERN, row_text):
            comp_id = f"DFR-{match.group(2)}"
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.DFR,
                    description=f"Digital fault recorder {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        # ── DFR modules ──
        for match in re.finditer(CM_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.DFR,
                    description=f"DFR current module {comp_id}",
                )

        for match in re.finditer(VM_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.DFR,
                    description=f"DFR voltage module {comp_id}",
                )

        # ── Custody meters ──
        for match in re.finditer(CMET_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.CUSTODY_METER,
                    description=f"Custody meter {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        # ── Fiber patch panels (normalize FPP2 → FPP-2) ──
        for match in re.finditer(FPP_NORM_PATTERN, row_text):
            comp_id = f"FPP-{match.group(2)}"
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.FIBER_PATCH,
                    description=f"Fiber patch panel {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        # ── Hand switches ──
        for match in re.finditer(HS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.HAND_SWITCH,
                    description=f"Hand switch {comp_id}",
                )

        # ── IRIG-B tees ──
        for match in re.finditer(TEE_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.UNKNOWN,
                    description=f"IRIG-B distribution tee {comp_id}",
                )

        # ── Auxiliary contacts ──
        for match in re.finditer(AUX_CONTACT_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.BREAKER,
                    description=f"Breaker auxiliary contact {comp_id}",
                )

        # ── Watt sensing links ──
        for match in re.finditer(WSL_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.CT,
                    description=f"Watt sensing link {comp_id}",
                )

        # ── Power supplies ──
        for match in re.finditer(PS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.POWER_SUPPLY,
                    description=f"Power supply {comp_id}",
                )

        # ── Relay panels ──
        for match in re.finditer(RP_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.PANEL,
                    description=f"Relay panel {comp_id}",
                )

        # ── Electronic trip modules ──
        for match in re.finditer(ETM_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RELAY,
                    description=f"Electronic trip module {comp_id}",
                )

        # ── Load tap changers ──
        for match in re.finditer(LTC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.LTC,
                    description=f"Load tap changer {comp_id}",
                )

        # ── Station service transformer ──
        for match in re.finditer(SST_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.TRANSFORMER,
                    description="Station service transformer",
                )

        # ── AC panels ──
        for match in re.finditer(AC_PANEL_PATTERN, row_text):
            comp_id = match.group(1).strip()
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.PANEL,
                    description=f"AC panel {comp_id}",
                )

        # ── Named fuses (FU-CL, FU-SST) ──
        for match in re.finditer(NAMED_FUSE_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.FUSE,
                    description=f"Fuse {comp_id}",
                )

        # ── BAF fuses ──
        for match in re.finditer(BAF_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.FUSE,
                    description=f"BAF fuse {comp_id}",
                )

        # ── BESS feeders ──
        for match in re.finditer(BESS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.BATTERY,
                    description=f"Battery energy storage feeder {comp_id}",
                )

        # ── IRIG-B time synchronization signal ──
        for match in re.finditer(IRIG_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.UNKNOWN,
                    description="IRIG-B time synchronization signal",
                )

        # ── Trip coils ──
        for match in re.finditer(TC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.BREAKER,
                    description=f"Trip coil {comp_id}",
                )

        # ── Lockout relay output ──
        for match in re.finditer(LOR_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.LOCKOUT,
                    description="Lockout relay output",
                )

        # ── DC voltage supply ──
        for match in re.finditer(VDC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.POWER_SUPPLY,
                    description=f"DC voltage supply {comp_id}",
                )

        # ── Communication modules ──
        for match in re.finditer(COMM_MODULE_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RELAY,
                    description=f"Communication module {comp_id}",
                )

        # ── CT class identifiers ──
        for match in re.finditer(CT_CLASS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.CT,
                    description=f"CT ratio class {comp_id}",
                )

        # ── Standalone terminal blocks ──
        for match in re.finditer(STANDALONE_TB_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.TERMINAL_BLOCK,
                    description=f"Terminal block {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        for match in re.finditer(STANDALONE_TS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.TERMINAL_BLOCK,
                    description=f"Test switch {comp_id}",
                )
            self._add_row_attributes(
                drawing.components[comp_id], cells, header,
            )

        # ── Substation automation & communication ──
        for match in re.finditer(RTAC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RTAC,
                    description="Real-time automation controller",
                )

        for match in re.finditer(CLOCK_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.GPS_CLOCK,
                    description="GPS clock / time synchronization device",
                )

        for match in re.finditer(PDC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.PDC,
                    description="Phasor data concentrator",
                )

        for match in re.finditer(CISCO_PATTERN, row_text):
            comp_id = match.group(1).strip()
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.NETWORK_SWITCH,
                    description=f"Network switch {comp_id}",
                )

        for match in re.finditer(RTR_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.ROUTER,
                    description=f"Network router {comp_id}",
                )

        for match in re.finditer(POWERSTRIP_PATTERN, row_text):
            comp_id = match.group(1).strip()
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.POWER_SUPPLY,
                    description="Power strip",
                )

        for match in re.finditer(ATS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.ATS,
                    description="Automatic transfer switch",
                )

        for match in re.finditer(RTU_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RTAC,
                    description="Remote terminal unit",
                )

        for match in re.finditer(HVAC_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.UNKNOWN,
                    description=f"HVAC unit {comp_id}",
                )

        for match in re.finditer(BP_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.PT,
                    description=f"Bushing potential device {comp_id}",
                )

        for match in re.finditer(EMS_PATTERN, row_text):
            comp_id = match.group(1)
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.RTAC,
                    description=f"Energy management system {comp_id}",
                )

        for match in re.finditer(BATTERY_EQUIP_PATTERN, row_text):
            comp_id = match.group(1).strip()
            if comp_id not in drawing.components:
                drawing.components[comp_id] = Component(
                    component_id=comp_id,
                    component_type=ComponentType.BATTERY,
                    description=comp_id,
                )

        # ── Future-proof equipment ──
        _future_patterns = [
            (UPS_PATTERN, ComponentType.UPS, "Uninterruptible power supply"),
            (MCC_PATTERN, ComponentType.MCC, "Motor control center"),
            (SWGR_PATTERN, ComponentType.SWGR, "Switchgear"),
            (VFD_PATTERN, ComponentType.VFD, "Variable frequency drive"),
            (PLC_PATTERN, ComponentType.PLC, "Programmable logic controller"),
            (RECTIFIER_PATTERN, ComponentType.RECTIFIER, "Rectifier"),
            (MOV_PATTERN, ComponentType.MOV, "Metal oxide varistor"),
            (ARRESTER_PATTERN, ComponentType.MOV, "Surge arrester"),
            (SPD_PATTERN, ComponentType.SPD, "Surge protective device"),
            (REGULATOR_PATTERN, ComponentType.REGULATOR, "Voltage regulator"),
            (GENERATOR_PATTERN, ComponentType.GENERATOR, "Generator"),
        ]
        for pattern, comp_type, desc_prefix in _future_patterns:
            for match in re.finditer(pattern, row_text):
                comp_id = match.group(1)
                if comp_id not in drawing.components:
                    drawing.components[comp_id] = Component(
                        component_id=comp_id,
                        component_type=comp_type,
                        description=f"{desc_prefix} {comp_id}",
                    )

        # ── Extract electrical values from row text ──
        for match in re.finditer(VOLTAGE_PATTERN, row_text):
            numeric = float(match.group(1))
            unit = match.group(2)
            ac_dc = match.group(3) or ""
            value_str = f"{match.group(1)}{unit} {ac_dc}".strip()
            cv = ComponentValue(
                parameter="voltage_rating",
                value=value_str,
                unit=unit,
                numeric_value=numeric if unit.upper() == "KV" else numeric / 1000,
            )
            # Associate with any component found in this row
            for comp_id in list(drawing.components.keys()):
                if comp_id in row_text:
                    existing = [
                        v.value for v in drawing.components[comp_id].values
                    ]
                    if cv.value not in existing:
                        drawing.components[comp_id].values.append(cv)
                    break

        # ── Extract cable specs ──
        for match in re.finditer(CABLE_PATTERN, row_text):
            cable_spec = match.group(1)
            if cable_spec not in drawing.cable_schedule:
                drawing.cable_schedule.append(cable_spec)

        for match in re.finditer(MULTIPAIR_CABLE_PATTERN, row_text):
            spec = match.group(1)
            if spec not in drawing.cable_schedule:
                drawing.cable_schedule.append(spec)

        # ── Extract cross-references ──
        for match in re.finditer(r'(NRE-E[CP]-\d{3}\.\d)', row_text):
            ref = match.group(1)
            if ref not in drawing.cross_references:
                drawing.cross_references.append(ref)

        for match in re.finditer(DRAWING_REF_PATTERN, row_text):
            ref = match.group(1)
            if ref not in drawing.cross_references:
                drawing.cross_references.append(ref)

        # ── Extract terminal block connections ──
        for match in re.finditer(r'\b(TB\d+)-(\d+)\b', row_text):
            tb_id = match.group(1)
            terminal = match.group(2)
            if tb_id not in drawing.terminal_blocks:
                drawing.terminal_blocks[tb_id] = []
            if terminal not in drawing.terminal_blocks[tb_id]:
                drawing.terminal_blocks[tb_id].append(terminal)

        for match in re.finditer(r'\b(TS\d+)-(\d+)\b', row_text):
            ts_id = match.group(1)
            terminal = match.group(2)
            tb_key = f"TS-{ts_id}"
            if tb_key not in drawing.terminal_blocks:
                drawing.terminal_blocks[tb_key] = []
            if terminal not in drawing.terminal_blocks[tb_key]:
                drawing.terminal_blocks[tb_key].append(terminal)

    # ─── Helper methods ───────────────────────────────────────────────

    @staticmethod
    def _build_column_map(header: list) -> dict:
        """Build a normalized column name → index map from header row."""
        col_map = {}
        for idx, cell in enumerate(header):
            upper = cell.upper().strip()
            if not upper:
                continue

            # Map to normalized keys
            if "ITEM" in upper:
                col_map["item"] = idx
            if "QTY" in upper or "QUANTITY" in upper:
                col_map["qty"] = idx
            if "CATALOG" in upper:
                col_map["catalog"] = idx
            if "MATERIAL" in upper and "NAMEPLATE" not in upper:
                col_map["material"] = idx
            if "DEVICE" in upper:
                col_map["device"] = idx
            if "DESCRIPTION" in upper or "NAME" in upper:
                col_map.setdefault("material", idx)
            if "SIZE" in upper and "LETTER" not in upper:
                col_map["size"] = idx
            if "LETTER" in upper:
                col_map["letter_size"] = idx
            if "1ST" in upper or "1ST LINE" in upper:
                col_map["1st_line"] = idx
            elif "LINE" in upper and "1" in upper:
                col_map["1st_line"] = idx
            if "2ND" in upper or "2ND LINE" in upper:
                col_map["2nd_line"] = idx
            elif "LINE" in upper and "2" in upper:
                col_map["2nd_line"] = idx

            # Always store raw column name too
            col_map[upper] = idx

        return col_map

    @staticmethod
    def _get_col(cells: list, col_map: dict, key: str,
                 default: str = "") -> str:
        """Get a cell value by normalized column key."""
        idx = col_map.get(key)
        if idx is not None and idx < len(cells):
            return cells[idx]
        return default

    @staticmethod
    def _add_row_attributes(comp: Component, cells: list, header: list):
        """Add cell values as component attributes using header as keys."""
        for i, cell in enumerate(cells):
            if not cell or cell == comp.component_id:
                continue
            key = header[i] if i < len(header) and header[i] else f"col_{i}"
            key = key.strip()
            if key and cell:
                comp.attributes[key] = cell

    @staticmethod
    def _extract_values_from_text(text: str, comp: Component):
        """Extract embedded electrical values from description text."""
        from .pdf_extractor import (
            VOLTAGE_PATTERN, CURRENT_PATTERN, IMPEDANCE_PATTERN,
        )

        for match in re.finditer(VOLTAGE_PATTERN, text):
            numeric = float(match.group(1))
            unit = match.group(2)
            ac_dc = match.group(3) or ""
            value_str = f"{match.group(1)}{unit} {ac_dc}".strip()
            comp.values.append(ComponentValue(
                parameter="voltage_rating",
                value=value_str,
                unit=unit,
                numeric_value=numeric if unit.upper() == "KV" else numeric / 1000,
            ))

        for match in re.finditer(CURRENT_PATTERN, text):
            numeric = float(match.group(1))
            unit = match.group(2)
            comp.values.append(ComponentValue(
                parameter="current_rating",
                value=f"{match.group(1)}{unit}",
                unit=unit,
                numeric_value=numeric,
            ))

        for match in re.finditer(IMPEDANCE_PATTERN, text):
            comp.values.append(ComponentValue(
                parameter="impedance",
                value=match.group(0),
                unit=match.group(2),
                numeric_value=float(match.group(1)),
            ))

    @staticmethod
    def _infer_type_from_designation(comp_id: str,
                                     description: str = "") -> ComponentType:
        """Infer ComponentType from a nameplate designation like TB1, FU1."""
        upper = comp_id.upper()
        desc_upper = description.upper() if description else ""

        if re.match(r'^(LT\d+-)?TB\d*$', upper):
            return ComponentType.TERMINAL_BLOCK
        if re.match(r'^(LT\d+-)?TS\d*$', upper):
            return ComponentType.TERMINAL_BLOCK
        if re.match(r'^FU\d*$', upper) or "FUSE" in desc_upper:
            return ComponentType.FUSE
        if re.match(r'^SW-?\d*$', upper):
            return ComponentType.SWITCH
        if "JUNCTION BOX" in desc_upper or "ENCLOSURE" in desc_upper:
            return ComponentType.JUNCTION_BOX
        if re.match(r'^CT-', upper):
            return ComponentType.CT
        if re.match(r'^VT-', upper):
            return ComponentType.VT
        if re.match(r'^PT-', upper):
            return ComponentType.PT
        if re.match(r'^CCVT', upper):
            return ComponentType.CCVT

        return ComponentType.BOM_ITEM

    @staticmethod
    def _categorize_cell(text: str, header: list, col_idx: int) -> str:
        """Categorize a cell value."""
        if re.search(r'\b(?:52|50|51|86|87|89|21|25|27|59|67|81)-[A-Z0-9]', text):
            return "component"
        if re.search(r'SEL-\d+|CT-|PT-|VT-|CCVT-', text):
            return "component"
        if re.search(r'FU\d+|BAF-\d+|TB\d+|TS\d+|SW-\d+', text):
            return "component"
        if re.search(r'RTAC|CLOCK|PDC|DFR|FPP|DPAC', text):
            return "component"
        if re.search(r'\d+\s*(?:kV|V|A|kA|MVA)', text):
            return "value"
        if re.search(r'\d+/C\s*#\d+', text):
            return "cable"

        # Use header to categorize
        if col_idx < len(header):
            h = header[col_idx].upper()
            if any(k in h for k in ["DRAWING", "DWG", "NUMBER"]):
                return "reference"
            if any(k in h for k in ["DESCRIPTION", "NAME", "TITLE", "MATERIAL"]):
                return "description"
            if any(k in h for k in ["QTY", "QUANTITY", "COUNT"]):
                return "quantity"
            if any(k in h for k in ["CATALOG", "PART"]):
                return "reference"

        return "text"
