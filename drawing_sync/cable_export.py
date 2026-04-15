"""Cable list extraction and Excel export.

Queries the connections table and cable schedule data from the database,
enriches with component attributes, and writes a formatted XLSX workbook
with three sheets: Cable List, Cable Schedule Summary, and By Drawing.
"""

import json
from datetime import datetime
from typing import Optional, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .db import ComponentDatabase


# ── Style constants ────────────────────────────────────────────────────
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_BOLD_FONT = Font(name="Calibri", bold=True, size=11)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


class CableListExporter:
    """Export cable/connection data from the database to a formatted Excel workbook."""

    def __init__(self, db: ComponentDatabase):
        self.db = db

    # ── Public API ─────────────────────────────────────────────────────

    def export_cable_list(
        self,
        output_path: str,
        filter_drawing: Optional[str] = None,
    ) -> dict:
        """Export cable list to an XLSX file.

        Args:
            output_path: Destination path for the workbook.
            filter_drawing: If provided, only include connections from this drawing.

        Returns:
            Summary dict with total_cables, unique_specs, drawings_covered, output_path.
        """
        cable_data = self._gather_cable_data(filter_drawing)
        schedule_data = self._gather_cable_schedule_data(filter_drawing)
        self._create_workbook(cable_data, schedule_data, output_path)

        unique_specs = set()
        unique_drawings = set()
        for row in cable_data:
            if row["cable_spec"]:
                unique_specs.add(row["cable_spec"])
            unique_drawings.add(row["source_drawing"])

        return {
            "total_cables": len(cable_data),
            "unique_specs": len(unique_specs),
            "drawings_covered": len(unique_drawings),
            "output_path": output_path,
        }

    # ── Data gathering ─────────────────────────────────────────────────

    def _gather_cable_data(self, filter_drawing: Optional[str] = None) -> List[dict]:
        """Query connections table and enrich with component attribute data."""
        if filter_drawing:
            rows = self.db.conn.execute(
                """
                SELECT c.drawing_id, c.from_component, c.from_terminal,
                       c.to_component, c.to_terminal,
                       c.cable_spec, c.wire_label, c.signal_type
                FROM connections c
                WHERE c.drawing_id = ?
                ORDER BY c.drawing_id, c.from_component, c.to_component
                """,
                (filter_drawing,),
            ).fetchall()
        else:
            rows = self.db.conn.execute(
                """
                SELECT c.drawing_id, c.from_component, c.from_terminal,
                       c.to_component, c.to_terminal,
                       c.cable_spec, c.wire_label, c.signal_type
                FROM connections c
                ORDER BY c.drawing_id, c.from_component, c.to_component
                """
            ).fetchall()

        # Build a lookup cache for component attributes (signals enrichment)
        signal_cache: Dict[str, str] = {}

        cable_data: List[dict] = []
        for r in rows:
            signal_type = r["signal_type"] or ""

            # Enrich: when signal_type is empty, look up from_component attributes
            if not signal_type:
                cache_key = f"{r['from_component']}||{r['drawing_id']}"
                if cache_key not in signal_cache:
                    signal_cache[cache_key] = self._lookup_signal(
                        r["from_component"], r["drawing_id"],
                    )
                signal_type = signal_cache[cache_key]

            # Generate cable number
            wire_label = r["wire_label"] or ""
            if wire_label:
                cable_number = wire_label
            else:
                cable_number = f"{r['from_component']}-{r['to_component']}"

            cable_data.append({
                "cable_number": cable_number,
                "from_component": r["from_component"],
                "from_terminal": r["from_terminal"] or "",
                "to_component": r["to_component"],
                "to_terminal": r["to_terminal"] or "",
                "cable_spec": r["cable_spec"] or "",
                "wire_label": wire_label,
                "signal_type": signal_type,
                "source_drawing": r["drawing_id"],
                "notes": "",
            })

        return cable_data

    def _lookup_signal(self, component_id: str, drawing_id: str) -> str:
        """Look up the 'signals' key in a component's attributes_json."""
        row = self.db.conn.execute(
            """
            SELECT attributes_json FROM components
            WHERE component_id = ? AND drawing_id = ?
            """,
            (component_id, drawing_id),
        ).fetchone()

        if row and row["attributes_json"]:
            try:
                attrs = json.loads(row["attributes_json"])
                signals = attrs.get("signals")
                if signals:
                    if isinstance(signals, list):
                        return ", ".join(signals)
                    return str(signals)
            except (json.JSONDecodeError, TypeError):
                pass
        return ""

    def _gather_cable_schedule_data(
        self, filter_drawing: Optional[str] = None,
    ) -> List[dict]:
        """Extract cable_schedule_json entries from the drawings table."""
        if filter_drawing:
            rows = self.db.conn.execute(
                """
                SELECT drawing_id, cable_schedule_json
                FROM drawings
                WHERE cable_schedule_json != '[]' AND drawing_id = ?
                ORDER BY drawing_id
                """,
                (filter_drawing,),
            ).fetchall()
        else:
            rows = self.db.conn.execute(
                """
                SELECT drawing_id, cable_schedule_json
                FROM drawings
                WHERE cable_schedule_json != '[]'
                ORDER BY drawing_id
                """
            ).fetchall()

        schedule_data: List[dict] = []
        for r in rows:
            try:
                specs = json.loads(r["cable_schedule_json"])
            except (json.JSONDecodeError, TypeError):
                continue
            for spec in specs:
                schedule_data.append({
                    "drawing_id": r["drawing_id"],
                    "cable_spec": spec if isinstance(spec, str) else str(spec),
                })

        return schedule_data

    # ── Workbook creation ──────────────────────────────────────────────

    def _create_workbook(
        self,
        cable_data: List[dict],
        schedule_data: List[dict],
        output_path: str,
    ) -> None:
        """Create the XLSX workbook with three sheets."""
        wb = Workbook()

        # ── Sheet 1: Cable List ────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Cable List"
        columns_1 = [
            "Cable Number",
            "From Component",
            "From Terminal",
            "To Component",
            "To Terminal",
            "Cable Type/Spec",
            "Wire Label",
            "Signal Type",
            "Source Drawing",
            "Notes",
        ]
        self._format_header(ws1, columns_1)

        for idx, row in enumerate(cable_data, start=2):
            ws1.cell(row=idx, column=1, value=row["cable_number"])
            ws1.cell(row=idx, column=2, value=row["from_component"])
            ws1.cell(row=idx, column=3, value=row["from_terminal"])
            ws1.cell(row=idx, column=4, value=row["to_component"])
            ws1.cell(row=idx, column=5, value=row["to_terminal"])
            ws1.cell(row=idx, column=6, value=row["cable_spec"])
            ws1.cell(row=idx, column=7, value=row["wire_label"])
            ws1.cell(row=idx, column=8, value=row["signal_type"])
            ws1.cell(row=idx, column=9, value=row["source_drawing"])
            ws1.cell(row=idx, column=10, value=row["notes"])

            # Alternating row colour
            if idx % 2 == 0:
                for col in range(1, len(columns_1) + 1):
                    ws1.cell(row=idx, column=col).fill = _ALT_ROW_FILL

        self._autofit_columns(ws1, len(columns_1))

        # ── Sheet 2: Cable Schedule Summary ────────────────────────────
        ws2 = wb.create_sheet("Cable Schedule Summary")
        columns_2 = [
            "Cable Spec",
            "Count",
            "Drawings Used In",
            "From Components",
            "To Components",
        ]
        self._format_header(ws2, columns_2)

        # Aggregate by cable_spec from the connections data
        spec_groups: Dict[str, dict] = {}
        for row in cable_data:
            spec = row["cable_spec"] or "(unspecified)"
            if spec not in spec_groups:
                spec_groups[spec] = {
                    "count": 0,
                    "drawings": set(),
                    "from_components": set(),
                    "to_components": set(),
                }
            spec_groups[spec]["count"] += 1
            spec_groups[spec]["drawings"].add(row["source_drawing"])
            spec_groups[spec]["from_components"].add(row["from_component"])
            spec_groups[spec]["to_components"].add(row["to_component"])

        # Also fold in schedule_data specs that might not appear in connections
        for sd in schedule_data:
            spec = sd["cable_spec"] or "(unspecified)"
            if spec not in spec_groups:
                spec_groups[spec] = {
                    "count": 0,
                    "drawings": set(),
                    "from_components": set(),
                    "to_components": set(),
                }
            spec_groups[spec]["drawings"].add(sd["drawing_id"])

        for idx, (spec, info) in enumerate(
            sorted(spec_groups.items(), key=lambda x: x[1]["count"], reverse=True),
            start=2,
        ):
            ws2.cell(row=idx, column=1, value=spec)
            ws2.cell(row=idx, column=2, value=info["count"])
            ws2.cell(row=idx, column=3, value=", ".join(sorted(info["drawings"])))
            ws2.cell(row=idx, column=4, value=", ".join(sorted(info["from_components"])))
            ws2.cell(row=idx, column=5, value=", ".join(sorted(info["to_components"])))

            if idx % 2 == 0:
                for col in range(1, len(columns_2) + 1):
                    ws2.cell(row=idx, column=col).fill = _ALT_ROW_FILL

        self._autofit_columns(ws2, len(columns_2))

        # ── Sheet 3: By Drawing ────────────────────────────────────────
        ws3 = wb.create_sheet("By Drawing")
        columns_3 = columns_1  # same columns as Cable List
        self._format_header(ws3, columns_3)

        # Group cable_data by source_drawing
        by_drawing: Dict[str, List[dict]] = {}
        for row in cable_data:
            by_drawing.setdefault(row["source_drawing"], []).append(row)

        current_row = 2
        for drawing_id in sorted(by_drawing.keys()):
            # Bold separator row with drawing name
            cell = ws3.cell(row=current_row, column=1, value=drawing_id)
            cell.font = _BOLD_FONT
            # Merge across all columns for the separator
            ws3.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(columns_3),
            )
            fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            for col in range(1, len(columns_3) + 1):
                ws3.cell(row=current_row, column=col).fill = fill
                ws3.cell(row=current_row, column=col).border = _THIN_BORDER
            current_row += 1

            for row in by_drawing[drawing_id]:
                ws3.cell(row=current_row, column=1, value=row["cable_number"])
                ws3.cell(row=current_row, column=2, value=row["from_component"])
                ws3.cell(row=current_row, column=3, value=row["from_terminal"])
                ws3.cell(row=current_row, column=4, value=row["to_component"])
                ws3.cell(row=current_row, column=5, value=row["to_terminal"])
                ws3.cell(row=current_row, column=6, value=row["cable_spec"])
                ws3.cell(row=current_row, column=7, value=row["wire_label"])
                ws3.cell(row=current_row, column=8, value=row["signal_type"])
                ws3.cell(row=current_row, column=9, value=row["source_drawing"])
                ws3.cell(row=current_row, column=10, value=row["notes"])

                if current_row % 2 == 0:
                    for col in range(1, len(columns_3) + 1):
                        ws3.cell(row=current_row, column=col).fill = _ALT_ROW_FILL

                current_row += 1

        self._autofit_columns(ws3, len(columns_3))

        # ── Save ───────────────────────────────────────────────────────
        import os
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)

    # ── Formatting helpers ─────────────────────────────────────────────

    def _format_header(self, ws, columns: list) -> None:
        """Write header row with blue fill / white bold text and freeze panes."""
        for col_idx, title in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGNMENT
            cell.border = Border(
                bottom=Side(style="medium", color="2F5496"),
            )

        # Freeze the row below the header
        ws.freeze_panes = "A2"

    @staticmethod
    def _autofit_columns(ws, num_columns: int) -> None:
        """Auto-fit column widths based on content, clamped to [10, 40]."""
        for col_idx in range(1, num_columns + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            # Add a little padding
            width = min(max(max_len + 2, 10), 40)
            ws.column_dimensions[col_letter].width = width
